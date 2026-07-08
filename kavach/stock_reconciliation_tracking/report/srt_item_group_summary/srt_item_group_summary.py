# Copyright (c) 2026, chaizup / kavach and contributors
# For license information, please see license.txt
# =============================================================================
# CONTEXT: SRT Item Group Summary — audit-coverage dashboard at ITEM GROUP
#   grain (user spec 2026-07-08, column structure from screenshot). One row
#   per Item Group: item-master counts, SRT activity in a CREATION-time
#   window, SRT approval-state buckets, and current stock valuation.
#
#   COLUMN SEMANTICS (documented decisions — see the report .md):
#     Total/Inactive/Active Items  → Item MASTER counts per group. Master
#       data: NOT affected by the date or warehouse filters.
#     Reconciliation (SRT) Count   → UNIQUE ITEMS of the group with ≥ 1
#       non-cancelled SRT (docstatus < 2 — draft, admin-approved,
#       super-admin-approved, system-approved ALL count) whose srt.creation
#       falls in [from_date, to_date] and whose default_warehouse matches
#       the warehouse filter. (User redefinition 2026-07-08: "how much uniq
#       item have response on srt record … except cancelled" — this is
#       COVERAGE in items, NOT a document count.)
#     Pending SRT Count Items      → items NOT FOUND in any such SRT record
#       ("not found in srt record according to date criteria") — the
#       un-audited gap: active_items − covered ACTIVE items. Disabled
#       items are excluded from the gap (a disabled item can't be counted;
#       leaving it in would inflate the backlog forever). With a warehouse
#       filter, "covered" means counted IN those warehouses.
#     System Approved (Matched)    → SRT docs at 'Approved By System'
#       (kavach Case 1 — every counted batch matched; no SR exists).
#     Admin Approval Pending       → drafts (docstatus = 0): the ADMIN has
#       not approved yet. (User correction 2026-07-08: this column is the
#       waiting-for-admin stage, NOT "got admin approval".)
#     Super Admin Approved Pending → docstatus = 1 at 'Admin Approval' (or
#       NULL/'' state — fail-pending): admin HAS approved, super admin
#       hasn't. DISJOINT from Admin Approval Pending by user spec — a doc
#       sits in exactly one stage; the two columns sum to the full waiting
#       pipeline.
#     Super Admin Approved Count   → docstatus = 1 at 'Super Admin
#       Approval': the super admin HAS approved (SR posted). Last column
#       (user 2026-07-08).
#   NOTE: the four stage columns count DOCUMENTS (their sum = total
#   non-cancelled SRT docs of the group in the window); srt_count and
#   pending_srt_items count ITEMS (coverage). Documents ≠ items whenever an
#   item was counted more than once in the window.
#     Total Stock Valuation        → SUM(Bin.stock_value) of the group's
#       items over the selected warehouses. CURRENT snapshot — the date
#       window does NOT rewind valuation (that needs an as-of SLE query;
#       out of scope per spec).
#
# INSTRUCTIONS:
#   - Date filter is on srt.creation (user: "from date to date filter of srt
#     CREATED time") — NOT posting_date. to_date is inclusive (< to+1day).
#   - all_time_pending (Check, "All Over Pending"): ON = Admin Approval
#     Pending + Super Admin Approved Pending ignore the date window (full
#     current backlog); every other column stays window-bound. The export
#     echo row flags it so a printed sheet can't be misread.
#   - Warehouse filter matches srt.default_warehouse for SRT columns and
#     bin.warehouse for valuation. Legacy SRTs with EMPTY default_warehouse
#     drop out when a warehouse filter is set (no way to attribute them).
#   - Rows appear for groups that have Items OR SRT activity in the window.
#
# RESTRICT:
#   - Bin.stock_value at ITEM level is fine here. The site's "never read
#     Bin" quirk applies to BATCH-level numbers only (batches live in the
#     SABB ledger) — do not "fix" this into a per-batch Bin read.
#   - Keep the pending predicate identical to srt_freeze.py's. If the SRT
#     workflow ever changes states, update BOTH (and kavach.md § Workflow).
# =============================================================================

import frappe
from frappe import _
from frappe.utils import add_days, cint, flt

# Waiting-pipeline predicate = admin_approval_pending ∪ super_admin_pending
# (docstatus-first, same rule as custom_erp_validation's srt_freeze.py —
# keep the two stage-column SUMs below in lock-step with it).


def execute(filters=None):
	filters = frappe._dict(filters or {})
	return get_columns(), get_data(filters)


def get_columns():
	return [
		{"fieldname": "item_group", "label": _("Item Group"), "fieldtype": "Link",
		 "options": "Item Group", "width": 180},
		{"fieldname": "total_items", "label": _("Total Count of Items"),
		 "fieldtype": "Int", "width": 130},
		{"fieldname": "inactive_items", "label": _("Inactive Items Count"),
		 "fieldtype": "Int", "width": 130},
		{"fieldname": "active_items", "label": _("Active Items Count"),
		 "fieldtype": "Int", "width": 130},
		{"fieldname": "srt_count", "label": _("Reconciliation (SRT) Count"),
		 "fieldtype": "Int", "width": 150},
		{"fieldname": "pending_srt_items", "label": _("Pending SRT Count Items"),
		 "fieldtype": "Int", "width": 150},
		{"fieldname": "system_approved", "label": _("System Approved (Matched)"),
		 "fieldtype": "Int", "width": 150},
		{"fieldname": "admin_approval_pending", "label": _("Admin Approval Pending"),
		 "fieldtype": "Int", "width": 140},
		{"fieldname": "super_admin_pending", "label": _("Super Admin Approved Pending"),
		 "fieldtype": "Int", "width": 170},
		{"fieldname": "stock_valuation", "label": _("Total Stock Valuation"),
		 "fieldtype": "Currency", "width": 160},
		{"fieldname": "super_admin_approved", "label": _("Super Admin Approved Count"),
		 "fieldtype": "Int", "width": 160},
	]


def get_data(filters):
	warehouses = _as_list(filters.get("warehouse"))
	item_groups = _as_list(filters.get("item_group"))

	# 1. Item master counts per group (master data — no date/warehouse scope)
	group_cond = "WHERE i.item_group IN %(groups)s" if item_groups else ""
	items = frappe.db.sql(
		f"""
		SELECT i.item_group,
		       COUNT(*)                          AS total_items,
		       SUM(i.disabled = 1)               AS inactive_items,
		       SUM(i.disabled = 0)               AS active_items
		FROM `tabItem` i
		{group_cond}
		GROUP BY i.item_group
		""",
		{"groups": tuple(item_groups) or ("",)},
		as_dict=True,
	)

	# 2. SRT activity + state buckets in the creation window
	conds = ["srt.docstatus < 2"]
	params = {"groups": tuple(item_groups) or ("",)}
	if filters.get("from_date"):
		conds.append("srt.creation >= %(from_date)s")
		params["from_date"] = filters.from_date
	if filters.get("to_date"):
		conds.append("srt.creation < %(to_date_excl)s")
		params["to_date_excl"] = add_days(filters.to_date, 1)
	if warehouses:
		conds.append("srt.default_warehouse IN %(warehouses)s")
		params["warehouses"] = tuple(warehouses)
	if item_groups:
		conds.append("i.item_group IN %(groups)s")

	srts = frappe.db.sql(
		f"""
		SELECT i.item_group,
		       COUNT(DISTINCT srt.item) AS srt_count,
		       COUNT(DISTINCT CASE WHEN i.disabled = 0 THEN srt.item END)
		           AS covered_active_items,
		       SUM(srt.docstatus = 1
		           AND srt.workflow_state = 'Approved By System')
		           AS system_approved,
		       SUM(srt.docstatus = 0)
		           AS admin_approval_pending,
		       SUM(srt.docstatus = 1
		           AND (srt.workflow_state = 'Admin Approval'
		                OR srt.workflow_state IS NULL
		                OR srt.workflow_state = ''))
		           AS super_admin_pending,
		       SUM(srt.docstatus = 1
		           AND srt.workflow_state = 'Super Admin Approval')
		           AS super_admin_approved
		FROM `tabStock Reconciliation SRT` srt
		JOIN `tabItem` i ON i.name = srt.item
		WHERE {" AND ".join(conds)}
		GROUP BY i.item_group
		""",
		params,
		as_dict=True,
	)

	# 3. Current stock valuation per group (warehouse-scoped, NOT date-scoped)
	val_conds = []
	if warehouses:
		val_conds.append("b.warehouse IN %(warehouses)s")
	if item_groups:
		val_conds.append("i.item_group IN %(groups)s")
	valuation = frappe.db.sql(
		f"""
		SELECT i.item_group, SUM(b.stock_value) AS stock_valuation
		FROM `tabBin` b
		JOIN `tabItem` i ON i.name = b.item_code
		{("WHERE " + " AND ".join(val_conds)) if val_conds else ""}
		GROUP BY i.item_group
		""",
		params,
		as_dict=True,
	)

	# Merge: any group with items OR SRT activity in the window
	by_group = {}
	for r in items:
		by_group[r.item_group] = frappe._dict(
			item_group=r.item_group,
			total_items=r.total_items, inactive_items=r.inactive_items,
			active_items=r.active_items,
			srt_count=0, covered_active_items=0, system_approved=0,
			admin_approval_pending=0, super_admin_pending=0,
			super_admin_approved=0, stock_valuation=0.0,
		)
	for r in srts:
		row = by_group.setdefault(r.item_group, frappe._dict(
			item_group=r.item_group, total_items=0, inactive_items=0,
			active_items=0, stock_valuation=0.0))
		row.update({k: r[k] for k in (
			"srt_count", "covered_active_items", "system_approved",
			"admin_approval_pending", "super_admin_pending",
			"super_admin_approved")})
	for r in valuation:
		if r.item_group in by_group:
			by_group[r.item_group].stock_valuation = flt(r.stock_valuation)

	# Pending = the un-audited gap: active items with NO non-cancelled SRT
	# in the window (disabled items excluded — see header).
	for row in by_group.values():
		covered = int(row.get("covered_active_items") or 0)
		row.pending_srt_items = max(0, int(row.get("active_items") or 0) - covered)

	# "All Over Pending" (user 2026-07-08): checkbox lifts the DATE window
	# from the two WAITING columns only — the true current backlog, however
	# old the SRT. Warehouse/group filters still apply. Safe to overwrite
	# only returned groups: window-pending ⊆ all-time-pending, so a group
	# absent here had zero in the window too.
	if cint(filters.get("all_time_pending")):
		at_conds = ["srt.docstatus < 2"]
		if warehouses:
			at_conds.append("srt.default_warehouse IN %(warehouses)s")
		if item_groups:
			at_conds.append("i.item_group IN %(groups)s")
		all_time = frappe.db.sql(
			f"""
			SELECT i.item_group,
			       SUM(srt.docstatus = 0) AS admin_approval_pending,
			       SUM(srt.docstatus = 1
			           AND (srt.workflow_state = 'Admin Approval'
			                OR srt.workflow_state IS NULL
			                OR srt.workflow_state = ''))
			           AS super_admin_pending
			FROM `tabStock Reconciliation SRT` srt
			JOIN `tabItem` i ON i.name = srt.item
			WHERE {" AND ".join(at_conds)}
			GROUP BY i.item_group
			""",
			params,
			as_dict=True,
		)
		for r in all_time:
			if r.item_group in by_group:
				by_group[r.item_group].admin_approval_pending = r.admin_approval_pending
				by_group[r.item_group].super_admin_pending = r.super_admin_pending

	return sorted(by_group.values(), key=lambda r: r.item_group or "")


def _as_list(value):
	"""MultiSelectList arrives as a JSON list; tolerate a comma string too."""
	if not value:
		return []
	if isinstance(value, str):
		import json
		try:
			return json.loads(value)
		except ValueError:
			return [v.strip() for v in value.split(",") if v.strip()]
	return list(value)


# =============================================================================
# export_xlsx — "Excel (Formatted)" button (house pattern cloned from
# chaizup_toc Purchase Batches Report). Frappe's stock Menu > Export drops
# styling entirely; this endpoint re-runs execute() with the SAME filters the
# user sees and builds a styled workbook:
#   row 1 title + row 2 filter echo · navy #0F172A header band (white bold,
#   wrapped) · zebra data rows · amber bold on non-zero pending columns ·
#   Int/valuation number formats · bold TOTAL row with top border (computed
#   here — execute() itself never returns a total row; the UI's total row
#   comes from add_total_row=1 and must stay OUT of the data) ·
#   freeze_panes just below the header · autofilter across it.
# RESTRICT:
#   - Keep openpyxl imports INSIDE the function — the report hot path must
#     not pay the import when nobody exports.
#   - frappe.response["type"] = "binary" + filecontent bytes; do NOT write
#     a File doc (no attachment litter for ad-hoc exports).
# =============================================================================
@frappe.whitelist()
def export_xlsx(filters=None):
	import io
	import json

	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	from frappe.utils import format_datetime, now_datetime

	filters = frappe._dict(json.loads(filters) if isinstance(filters, str) else (filters or {}))
	columns = get_columns()
	rows = get_data(filters)

	wb = Workbook()
	ws = wb.active
	ws.title = "SRT Item Group Summary"

	navy = "0F172A"
	zebra = "F1F5F9"
	amber = "B45309"

	# Row 1-2: title + filter echo
	ws.cell(row=1, column=1, value="SRT Item Group Summary").font = Font(
		bold=True, size=14, color=navy)
	wh = ", ".join(_as_list(filters.get("warehouse"))) or "All warehouses"
	ig = ", ".join(_as_list(filters.get("item_group"))) or "All item groups"
	pending_mode = (
		"  |  PENDING COLUMNS = ALL-TIME BACKLOG (date window ignored)"
		if cint(filters.get("all_time_pending")) else ""
	)
	ws.cell(row=2, column=1, value=(
		f"SRT created {filters.get('from_date') or '…'} → {filters.get('to_date') or '…'}"
		f"  |  {wh}  |  {ig}{pending_mode}"
		f"  |  generated {format_datetime(now_datetime())}"
	)).font = Font(size=9, color="64748B")

	# Row 4: header band
	head_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
	HEADER_ROW = 4
	for c_i, col in enumerate(columns, start=1):
		cell = ws.cell(row=HEADER_ROW, column=c_i, value=col["label"])
		cell.fill = head_fill
		cell.font = Font(bold=True, size=10, color="FFFFFF")
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		ws.column_dimensions[get_column_letter(c_i)].width = max(14, int(col["width"] / 7))
	ws.row_dimensions[HEADER_ROW].height = 30

	int_cols = {c["fieldname"] for c in columns if c["fieldtype"] == "Int"}
	pending_cols = {"pending_srt_items", "admin_approval_pending", "super_admin_pending"}
	totals = {f: 0 for f in int_cols} | {"stock_valuation": 0.0}

	r_i = HEADER_ROW
	for row in rows:
		r_i += 1
		stripe = PatternFill(start_color=zebra, end_color=zebra, fill_type="solid") \
			if (r_i - HEADER_ROW) % 2 == 0 else None
		for c_i, col in enumerate(columns, start=1):
			f = col["fieldname"]
			val = row.get(f)
			cell = ws.cell(row=r_i, column=c_i, value=val)
			if stripe:
				cell.fill = stripe
			if f in int_cols:
				cell.value = int(val or 0)
				cell.number_format = "#,##0"
				totals[f] += int(val or 0)
				if f in pending_cols and cell.value:
					cell.font = Font(bold=True, color=amber)
			elif f == "stock_valuation":
				cell.value = flt(val)
				cell.number_format = "#,##0.00"
				totals[f] += flt(val)

	# TOTAL row (bold, top border)
	r_i += 1
	top = Border(top=Side(style="medium", color=navy))
	for c_i, col in enumerate(columns, start=1):
		f = col["fieldname"]
		val = "TOTAL" if c_i == 1 else (totals.get(f) if f in totals else None)
		cell = ws.cell(row=r_i, column=c_i, value=val)
		cell.font = Font(bold=True, color=navy)
		cell.border = top
		if f in int_cols:
			cell.number_format = "#,##0"
		elif f == "stock_valuation":
			cell.number_format = "#,##0.00"

	ws.freeze_panes = f"A{HEADER_ROW + 1}"
	ws.auto_filter.ref = (
		f"A{HEADER_ROW}:{get_column_letter(len(columns))}{r_i - 1}"
	)

	buf = io.BytesIO()
	wb.save(buf)
	frappe.response["filename"] = (
		f"SRT_Item_Group_Summary_{filters.get('from_date') or 'all'}"
		f"_{filters.get('to_date') or 'all'}.xlsx"
	)
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"] = "binary"
